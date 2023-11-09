import openpyxl
from sys import argv

months = ['ינואר','פברואר','מרץ','אפריל','מאי','יוני','יולי','אוגוסט','ספטמבר','אוקטובר','נובמבר','דצמבר']


base_folder = "C:\\Shared Folder\\לקוחות\\א.ג.מ אילת נכסים בעמ"

if len(argv) < 3:
	raise SystemExit("python agm.py <year> <month>")

year = argv[1]
month = int(argv[2])

folder = f'{base_folder}\\הכנסות {year}\\הכנסות {month}.{year}'

files = [
	(f"{folder}\\אקסל חשבוניות וילות {months[month-1]}.xlsx","חשבוניות וילות"),
	(f"{folder}\\אקסל חשבוניות רויאל פארק {months[month-1]}.xlsx","חשבוניות רויאל"),
	(f"{folder}\\אקסל קבלות וילות {months[month-1]}.xlsx","קבלות וילות"),
	(f"{folder}\\אקסל קבלות רויאל פארק {months[month-1]}.xlsx",	"קבלות רויאל")
]


out_workbook = openpyxl.Workbook()

out_sheet = out_workbook.active
out_sheet.display_in_rtl = True
output_row_index = 0

def parse_file(filename,type,output_file,totals):
	global output_row_index
	income_type_array = None
	doc_type = None
	if type == "חשבוניות וילות":
		income_type_array = [160,66,5]
		doc_type = 'invoice'
	elif type == "חשבוניות רויאל":
		income_type_array = [150,66,6]
		doc_type = 'invoice royal'

	in_workbook = openpyxl.load_workbook(filename)
	in_sheet = in_workbook.active
	for row in in_sheet.iter_rows(values_only=True):
		document_date = row[0]
		if not row[0] or row[8] == 0 or row[0] == "תאריך":
			continue
		if row[0] == "סך הכל" or row[0] == "סוג":
			break
		if "חשבוניות" in type:
			row_values = [1,'',float(row[8]),float(row[10]),row[0],row[0],row[6],row[6],row[4]]
			row_values = income_type_array+row_values
		elif "קבלות" in type:
			date = row[0].split()[0]
			row_values = [1,'',float(row[8]),'',date,date,row[4],row[4],row[5]]
			row_values_dict = {
				"BIT" : [10,9,66]+row_values,
				"CASH" : [12,8,66]+row_values,
				"CC" : [14,11,66]+row_values,
				"NOCC" : [14,11,66]+row_values,
				"TR" : [10,9,66]+row_values,
				"CH" : [9,7,66]+row_values,
	
			}
			payment_type = row[6]
			if payment_type == "סוג תשלום":
				continue
			row_values = row_values_dict[payment_type]
			doc_type = payment_type
				
		row_values = [*map(str,row_values)]
		output_file.write('\t'.join(row_values) + '\n')
		if doc_type not in totals:
			totals[doc_type] = 0
		totals[doc_type] += float(row_values[5])
			
	

with open('טעינה לרווחית.txt','w') as f:
	totals = {}
	for file,type in files:
		parse_file(file,type,f,totals)
	print(totals)





		
		
	


